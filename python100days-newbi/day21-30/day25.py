import openpyxl as oxl
import random as rnd

# write to excel
def createWorkBook(file):
    # 1.create workbook
    workbook = oxl.Workbook()
    # 2 create a sheet
    sheet = workbook.active
    # 3.add title to the sheet
    sheet.title = "Final Score"
    # 4.table header
    headers = ('Name','Math','Chinese','English')
    # 5.write the header data
    for idx,val in enumerate(headers):
        sheet.cell(1,idx+1,val)
    # 6.names
    names = ('Jack','Mary','Jesse','Linda','Lili')
    for row,val in enumerate(names):
        sheet.cell(row+2,1,val)
        for i in range(2,5):
            sheet.cell(row+2,i,rnd.randrange(60,101))

    workbook.save(file)
    workbook.close()

# 2.read excel
def readExcel(file):
    # 2.1 load the workbook
    wb = oxl.load_workbook(file)
    # 2.2 get the sheet
    sheet = wb.active
    # print(sheet)
    # for line in sheet:
    #     print(line)
    for row in sheet.iter_rows(values_only=True):
        # print(row)
        for el in row:
            print(el,end="\t")
        print()   

    wb.close()

# readExcel("./score.xlsx")

# createWorkBook("score.xlsx")

